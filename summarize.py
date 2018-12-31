import collections
import argparse
import yaml
import re
import os
import util
import datetime
import openpyxl

# stu_id is array
Student = collections.namedtuple("Student", "stu_id name user_id")
# pass_info is map user_id: pass_num
Contest = collections.namedtuple("Contest", "id name begin prob_ids prob_num pass_info")
# contests is array sorted by start_time, student is array, user_stu is map
Semester = collections.namedtuple("Semester", "name contests students user_stu")


def get_config():
    parser = argparse.ArgumentParser(description='Detect cheat in vjudge contest.')
    # basic configuration
    parser.add_argument('-c', '--config', dest='config_file', help='YAML config file path')

    config = parser.parse_args()
    if config.config_file is not None:
        with open(config.config_file, "r", encoding='utf-8') as f:
            y = yaml.load(f)
            config.ignore_userid = util.ensure_list(y.get("ignore_userid", []))
            config.include_userid = util.ensure_list(y.get("include_userid", []))
            if len(config.include_userid) == 0:
                config.include_userid.append(".+")
            config.include_userid = list(map(lambda x: re.compile(f"^{x}$"), config.include_userid))
            config.semester = y.get("semester")
    print(config)
    return config

def build_students(base_dir, stu_info_path):
    stu_info_path = os.path.join(base_dir, stu_info_path)
    students = []
    with open(stu_info_path, 'r', encoding='utf-8') as f:
        for line in f:
            items = line.strip().split()
            if len(items) < 2:
                continue
            
            stu_id, name, *user_id = items
            if len(user_id) == 0:
                user_id.append(stu_id)
            stu = Student(stu_id, name, user_id)
            students.append(stu)
    return students

def build_user_stu(students):
    user_stu = {}
    for stu in students:
        for user_id in stu.user_id:
            assert(user_id not in user_stu)
            user_stu[user_id] = stu.stu_id
    return user_stu

def build_contests(base_dir, all_contests, students, user_stu):
    contests = []
    for zip_name in all_contests:
        cid = util.get_contest_id(zip_name)
        zip_path = os.path.join(base_dir, zip_name)
        assert(os.path.exists(zip_path))

        info = util.get_contest_info(cid)
        assert(info is not None)

        name = str(info["title"])
        begin_unix = int(info["begin"] / 1000)
        begin = datetime.datetime.fromtimestamp(begin_unix)

        unzip_dir = os.path.join(base_dir, name, "submissions")
        assert(os.path.exists(unzip_dir))

        prob_ids = os.listdir(unzip_dir)
        prob_num = len(prob_ids)

        pass_info = {}
        for stu in students:
            pass_info[stu.stu_id] = set()

        for problem in os.listdir(unzip_dir):
            problem_dir = os.path.join(unzip_dir, problem)
            for sub in os.listdir(problem_dir):
                user_id = util.get_userid(sub)
                if user_id not in user_stu:
                    continue
                stu_id = user_stu[user_id]
                pass_info[stu_id].add(problem)
        contests.append(Contest(cid, name, begin, prob_ids, prob_num, pass_info))
    return contests

def build_semester(config, base_dir):
    name = config.semester["name"]
    stu_info_path = config.semester["stu_info_path"]
    all_contests = config.semester["all_contests"]
    
    students = build_students(base_dir, stu_info_path)

    user_stu = build_user_stu(students)

    contests = build_contests(base_dir, all_contests, students, user_stu)
    contests = sorted(contests, key=lambda x: x.begin)
    semester = Semester(name, contests, students, user_stu)
    return semester

def output_semester(output_path, semester):
    book = openpyxl.Workbook(0)
    cols = len(semester.contests) + 5

    sheet = book.active
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    title = sheet.cell(1, 1)
    title.value = semester.name
    title.alignment = openpyxl.styles.Alignment(horizontal='center')
    title.font = openpyxl.styles.Font(bold=True, size=14)

    rows = []
    title_info = ['编号', '学号', '姓名', '用户名']
    total_probs = 0
    for c in semester.contests:
        title_info.append(c.begin.strftime("%m/%d") + f" ({c.prob_num})")
        total_probs += c.prob_num
    title_info.append(f'总和 ({total_probs})')
    rows.append(title_info)

    for i, stu in enumerate(semester.students):
        cur_row = [str(i + 1), stu.stu_id, stu.name]
        cur_row.append(str(stu.user_id) if len(stu.user_id) > 1 else stu.user_id[0])
        total_pass = 0
        for c in semester.contests:
            cur_pass = len(c.pass_info[stu.stu_id])
            cur_row.append(cur_pass)
            total_pass += cur_pass
        cur_row.append(total_pass)
        rows.append(cur_row)
        
    for row in rows:
        sheet.append(row)

    column_widths = []
    for row in rows:
        for i, cell in enumerate(row):
            cell = str(cell).encode('utf-8')
            cur_len = len(cell) + 2
            if len(column_widths) > i:
                if cur_len > column_widths[i]:
                    column_widths[i] = cur_len
            else:
                column_widths += [cur_len]

    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[chr(ord('A') + i)].width = column_width
    book.save(output_path)

def process(config):
    base_dir = os.path.dirname(config.config_file)
    semester = build_semester(config, base_dir)
    cur_time = datetime.datetime.now()
    cur_time_str = cur_time.strftime("%Y-%m-%d")
    output_path = os.path.join(base_dir, f"{semester.name}_{cur_time_str}.xlsx")
    print(output_path)
    output_semester(output_path, semester)

if __name__ == "__main__":
    process(get_config())
